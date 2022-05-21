from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

ws1 = wb.create_sheet("Mysheet") # insert at the end (default)

ws1 = wb['Mysheet']
# Data can be assigned directly to cells

cells = ws1['A1':'J10']

# 変数iの初期値を0にします。
i = 1

for row in cells:
    # for文を利用して、変数rowがある限り、変数cellに1個づつ代入します。
    for cell in row:
        # 変数cellの値に、変数iを代入します。
        cell.value = i
        # 変数cellの値に、変数iを代入します。
        i += 1

ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")